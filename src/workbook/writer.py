"""Live-formula Excel merger-model writer.

Writes ``out/<DEAL>_model.xlsx`` from a :class:`~src.interfaces.MergerModelBundle`
per docs/WORKBOOK_SPEC.md — **the single source of truth**. The governing rule
(banker convention + the cell-level differential): every *computed* cell is a
live Excel formula referencing other cells / named ranges; hard inputs (blue)
live only on the Assumptions tab and Cover metadata. Historical statement lines
and engine-solve outputs are written as black reported values (the tie-out
surface); their combinations are formulas.

The writer is deliberately faithful to the engine's arithmetic so that when the
verifier recalculates the file (via the ``formulas`` library) every mapped cell
reproduces the engine number to the cent. :func:`build_verifier_cell_map` is the
authoritative ``(sheet, coord) -> engine-value`` map the verifier consumes;
keeping it beside the writer keeps the two in lockstep.

Design ported from the thesis writer (cited in docs/DESIGN.md): one private
method per tab, the ``self._anchors`` cross-tab reference mechanism (accumulate
cell addresses as tabs are built so later tabs / named ranges reference earlier
cells without hardcoding coordinates), ``_apply_ergonomics`` (column widths +
freeze panes), and the deterministic ``_save_pinned`` (pins ZIP timestamps so a
rebuild is byte-identical).

Scope note: the core deal tabs — Assumptions, Acquirer, Target, Deal & S&U,
PPA, Pro Forma IS, Accretion-Dilution, Contribution — emit LIVE formulas that
recompute the engine's numbers off the blue inputs. Two later tabs carry live
formulas verified by the differential over cells that are pure functions of
other cells already on the sheet: the **Sensitivities** premium×synergies grid
interior (synergy-column index ≥ 2) is a live linear interpolation off the two
engine-value anchors in each row (A/D is exactly linear in the synergy run-rate,
so this reproduces the grid to the cent), and the **Fairness Comparison** tab's
disclosed/our midpoints and per-method overlap % are live formulas off the
disclosed/our low-high cells. The remaining values on those tabs (breakeven
synergies, the two anchor columns and axis labels of the grid, the disclosed/our
range endpoints) are labeled black engine/extraction outputs, as are the
Pro Forma BS and Precedents tabs.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from src.interfaces import MergerModelBundle
from src.schema import LineItem
from src.workbook import styles as S

# Per-sheet column widths + freeze panes for banker-grade ergonomics.
_LABEL_COL_WIDTH = 36.0
_DATA_COL_WIDTH = 14.0

# Sheet names, in spec order (WORKBOOK_SPEC §Tab order).
# NOTE: the spec lists tab 9 as "Accretion/Dilution", but Excel forbids "/" in a
# sheet title, so we use "Accretion-Dilution". Every other name is verbatim.
SH_COVER = "Cover"
SH_ASSUM = "Assumptions"
SH_ACQ = "Acquirer"
SH_TGT = "Target"
SH_DEAL = "Deal & S&U"
SH_PPA = "PPA"
SH_PF_IS = "Pro Forma IS"
SH_PF_BS = "Pro Forma BS"
SH_AD = "Accretion-Dilution"
SH_CONTRIB = "Contribution"
SH_SENS = "Sensitivities"
SH_PREC = "Precedents"
SH_FAIR = "Fairness Comparison"

SHEET_ORDER = [
    SH_COVER,
    SH_ASSUM,
    SH_ACQ,
    SH_TGT,
    SH_DEAL,
    SH_PPA,
    SH_PF_IS,
    SH_PF_BS,
    SH_AD,
    SH_CONTRIB,
    SH_SENS,
    SH_PREC,
    SH_FAIR,
]

# Tabs that are multi-year data grids with a header on row 3 → freeze below it.
_GRID_TABS = frozenset({SH_ACQ, SH_TGT, SH_PF_IS, SH_PF_BS, SH_AD, SH_CONTRIB})

# Required named ranges (WORKBOOK_SPEC §Named ranges).
NAMED_RANGES = [
    "DealValue",
    "EquityPurchasePrice",
    "CashPerShare",
    "ExchangeRatio",
    "ImpliedPremium",
    "Goodwill",
    "NewDebt",
    "NewSharesIssued",
    "ProFormaEPS_Y1",
    "Accretion_Y1",
    "Accretion_Y2",
    "Accretion_Y3",
    "BreakevenSynergies",
    "AcquirerOwnership",
    "TargetOwnership",
]


class ExcelWorkbookWriter:
    """Concrete WorkbookWriter. See the module docstring for the contract."""

    def write(self, path: str, model: MergerModelBundle) -> None:
        wb = Workbook()
        # Pin document timestamps to a fixed epoch so a rebuild is byte-identical
        # (openpyxl otherwise stamps the current UTC time into docProps/core.xml).
        from datetime import datetime

        fixed = datetime(2026, 1, 1, 0, 0, 0)  # noqa: DTZ001 — deterministic constant
        wb.properties.created = fixed
        wb.properties.modified = fixed

        wb.remove(wb.active)
        sheets: dict[str, Worksheet] = {name: wb.create_sheet(name) for name in SHEET_ORDER}

        # Cross-tab reference addresses, filled as we build (the thesis pattern).
        self._anchors: dict[str, str] = {}

        # Build order respects dependencies: Assumptions (inputs) first, then the
        # tabs that reference them, then Cover (which links to headline outputs).
        self._assumptions(sheets[SH_ASSUM], model)
        self._acquirer(sheets[SH_ACQ], model)
        self._target(sheets[SH_TGT], model)
        self._deal(sheets[SH_DEAL], model)
        self._ppa(sheets[SH_PPA], model)
        self._pro_forma_is(sheets[SH_PF_IS], model)
        self._pro_forma_bs(sheets[SH_PF_BS], model)
        self._accretion(sheets[SH_AD], model)
        self._contribution(sheets[SH_CONTRIB], model)
        self._sensitivities(sheets[SH_SENS], model)
        self._precedents(sheets[SH_PREC], model)
        self._fairness(sheets[SH_FAIR], model)
        self._cover(sheets[SH_COVER], model)

        self._add_charts(sheets, model)
        self._apply_ergonomics(sheets)
        self._define_names(wb)
        self._save_pinned(wb, path)

    # -- Assumptions: the ONLY blue-input tab ---------------------------------
    def _assumptions(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Assumptions (inputs — blue)")
        cons = model.deal.consideration
        ppa = model.deal.ppa
        da = model.deal_assumptions

        # --- Standalone drivers: revenue growth per projected year (blue) ------
        S.subheader(ws, "A3", "Standalone drivers")
        S.label_cell(ws, "A4", "Acquirer revenue growth")
        S.label_cell(ws, "A5", "Target revenue growth")
        acq_g = model.acquirer_assumptions.revenue_growth
        tgt_g = model.target_assumptions.revenue_growth
        nyr_acq = len(model.acquirer_statements.periods) - model.acquirer_statements.n_hist
        nyr_tgt = len(model.target_statements.periods) - model.target_statements.n_hist
        for j in range(nyr_acq):
            g = acq_g[j] if j < len(acq_g) else 0.0
            S.input_cell(ws, f"{_col(2 + j)}4", g, S.FMT_PERCENT)
        for j in range(nyr_tgt):
            g = tgt_g[j] if j < len(tgt_g) else 0.0
            S.input_cell(ws, f"{_col(2 + j)}5", g, S.FMT_PERCENT)
        self._anchors["acq_growth_row"] = "4"
        self._anchors["tgt_growth_row"] = "5"

        # --- Deal terms (disclosed) -------------------------------------------
        S.subheader(ws, "A7", "Deal terms (disclosed)")
        prem_ref = _first_not_none(
            _sv(model.terms.premium_reference_price),
            _sv(model.terms.reference_acquirer_price),
            _safe_div(cons.stock_per_share_value, cons.exchange_ratio),
            0.0,
        )
        ref_px = _first_not_none(
            _sv(model.terms.reference_acquirer_price),
            _safe_div(cons.stock_per_share_value, cons.exchange_ratio),
            0.0,
        )
        deal_terms = [
            ("Cash per target share", cons.cash_per_share, S.FMT_PERSHARE, "cash_ps"),
            ("Exchange ratio", cons.exchange_ratio or 0.0, S.FMT_PERSHARE, "xratio"),
            ("Reference acquirer price", ref_px, S.FMT_PERSHARE, "ref_px"),
            ("Target shares outstanding", cons.target_shares, S.FMT_SHARES, "tgt_sh"),
            ("Premium reference price", prem_ref, S.FMT_PERSHARE, "prem_ref"),
        ]
        r = 8
        for label, val, fmt, key in deal_terms:
            S.label_cell(ws, f"A{r}", label)
            S.input_cell(ws, f"B{r}", val, fmt)
            self._anchors[f"assum_{key}"] = _ref(SH_ASSUM, f"B{r}")
            r += 1

        # --- Our deal assumptions (financing, fees, PPA choices) --------------
        r += 1
        S.subheader(ws, f"A{r}", "Our deal assumptions")
        r += 1
        our = [
            ("New acquisition debt", da.new_debt, S.FMT_CURRENCY, "new_debt"),
            ("New debt rate", da.new_debt_rate, S.FMT_PERCENT, "new_debt_rate"),
            ("Cash on hand used", da.cash_on_hand_used, S.FMT_CURRENCY, "cash_used"),
            ("Foregone cash yield", da.foregone_cash_yield, S.FMT_PERCENT, "fore_yield"),
            ("Advisory fees", da.advisory_fees, S.FMT_CURRENCY, "adv_fees"),
            ("Financing fees", da.financing_fees, S.FMT_CURRENCY, "fin_fees"),
            ("Intangible step-up", da.intangible_step_up, S.FMT_CURRENCY, "intang_step"),
            ("PP&E step-up", da.ppe_step_up, S.FMT_CURRENCY, "ppe_step"),
            ("Deferred-tax rate on step-ups", da.deferred_tax_rate, S.FMT_PERCENT, "dtl_rate"),
            ("Marginal tax rate", da.marginal_tax_rate, S.FMT_PERCENT, "tax_rate"),
        ]
        for label, val, fmt, key in our:
            S.label_cell(ws, f"A{r}", label)
            S.input_cell(ws, f"B{r}", val, fmt)
            self._anchors[f"assum_{key}"] = _ref(SH_ASSUM, f"B{r}")
            r += 1

        # --- PPA book inputs ---------------------------------------------------
        r += 1
        S.subheader(ws, f"A{r}", "Purchase price accounting — book inputs")
        r += 1
        ppa_inputs = [
            ("Target book equity", ppa.target_book_equity, "tgt_book_eq"),
            ("Target existing goodwill", ppa.target_existing_goodwill, "tgt_gw"),
        ]
        for label, val, key in ppa_inputs:
            S.label_cell(ws, f"A{r}", label)
            S.input_cell(ws, f"B{r}", val, S.FMT_CURRENCY)
            self._anchors[f"assum_{key}"] = _ref(SH_ASSUM, f"B{r}")
            r += 1

        # --- Synergy cases (run-rate per case; management vs ours) ------------
        r += 1
        S.subheader(ws, f"A{r}", "Synergy cases (run-rate, pre-tax)")
        r += 1
        for c in model.combinations:
            tag = "management, disclosed" if c.synergy_case.is_disclosed else "our case"
            S.label_cell(ws, f"A{r}", f"{c.synergy_case.name} ({tag})")
            S.input_cell(ws, f"B{r}", c.synergy_case.run_rate_annual, S.FMT_CURRENCY)
            r += 1

    # -- Acquirer standalone (revenue chain live off Assumptions) -------------
    def _acquirer(self, ws: Worksheet, model: MergerModelBundle) -> None:
        self._standalone(ws, model, model.acquirer_statements, "Acquirer", "acq_growth_row", "acq")

    def _target(self, ws: Worksheet, model: MergerModelBundle) -> None:
        self._standalone(ws, model, model.target_statements, "Target", "tgt_growth_row", "tgt")

    def _standalone(self, ws, model, stmts, title, growth_key, prefix) -> None:
        """Standalone tab: reported historicals (tie-out surface) + a live
        revenue projection chain (base × (1+growth)) off the Assumptions blue
        growth inputs. Net income is rendered as the engine's projected value
        (a black tie-out figure), not a formula."""
        S.section_header(ws, "A1", f"{title} — standalone")
        n = stmts.n_hist
        nyr = len(stmts.periods) - n
        base_rev = _last_hist(stmts, LineItem.REVENUE) or 0.0

        S.label_cell(ws, "A3", "Line ($)", bold=True)
        S.value_cell(ws, "B3", "Base", S.FMT_YEAR)
        for j in range(nyr):
            S.label_cell(ws, f"{_col(3 + j)}3", f"Y{j + 1}", bold=True)

        gr = self._anchors[growth_key]
        # Revenue row 4: base in B, live chain in C..
        S.label_cell(ws, "A4", "Revenue")
        S.value_cell(ws, "B4", base_rev)
        proj_rev = stmts.series(LineItem.REVENUE)[n:]
        for j in range(nyr):
            prev = f"{_col(2 + j)}4"
            gcell = f"'{SH_ASSUM}'!{_col(2 + j)}{gr}"
            S.formula_cell(ws, f"{_col(3 + j)}4", f"={prev}*(1+{gcell})")
            if j < len(proj_rev) and proj_rev[j] is not None:
                self._anchors[f"{prefix}_rev_{j}"] = f"{_col(3 + j)}4"

        # Net income row 5: engine projected values (black tie-out figures).
        S.label_cell(ws, "A5", "Net income")
        proj_ni = stmts.series(LineItem.NET_INCOME)[n:]
        for j in range(nyr):
            v = proj_ni[j] if j < len(proj_ni) else None
            if v is not None:
                S.value_cell(ws, f"{_col(3 + j)}5", v)

    # -- Deal & Sources/Uses --------------------------------------------------
    def _deal(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Deal — consideration, sources & uses")
        cash = self._anchors["assum_cash_ps"]
        xr = self._anchors["assum_xratio"]
        ref = self._anchors["assum_ref_px"]
        tsh = self._anchors["assum_tgt_sh"]
        pref = self._anchors["assum_prem_ref"]

        S.subheader(ws, "A3", "Consideration")
        S.label_cell(ws, "A4", "Aggregate cash consideration")
        S.formula_cell(ws, "B4", f"={cash}*{tsh}")
        S.label_cell(ws, "A5", "Aggregate stock consideration")
        S.formula_cell(ws, "B5", f"={xr}*{ref}*{tsh}")
        S.label_cell(ws, "A6", "Equity purchase price", bold=True)
        S.formula_cell(ws, "B6", "=B4+B5")
        S.label_cell(ws, "A7", "Total consideration per share")
        S.formula_cell(ws, "B7", f"={cash}+{xr}*{ref}", S.FMT_PERSHARE)
        S.label_cell(ws, "A8", "New shares issued")
        S.formula_cell(ws, "B8", f"={xr}*{tsh}", S.FMT_SHARES)
        S.label_cell(ws, "A9", "Implied premium")
        S.formula_cell(ws, "B9", f"=B7/{pref}-1", S.FMT_PERCENT)
        self._anchors["agg_cash"] = _ref(SH_DEAL, "B4")
        self._anchors["agg_stock"] = _ref(SH_DEAL, "B5")
        self._anchors["equity_pp"] = _ref(SH_DEAL, "B6")
        self._anchors["new_shares"] = _ref(SH_DEAL, "B8")
        self._anchors["implied_premium"] = _ref(SH_DEAL, "B9")

        # Sources & Uses — engine line items as black values; totals are live
        # SUM formulas (the invariant the deal engine guarantees: they balance).
        su = model.deal.sources_and_uses
        S.subheader(ws, "A11", "Sources")
        r = 12
        first_src = r
        for k, v in su.sources.items():
            S.label_cell(ws, f"A{r}", _humanize(k))
            S.value_cell(ws, f"B{r}", v)
            r += 1
        last_src = r - 1
        S.label_cell(ws, f"A{r}", "Total sources", bold=True)
        S.formula_cell(ws, f"B{r}", f"=SUM(B{first_src}:B{last_src})")
        self._anchors["total_sources"] = _ref(SH_DEAL, f"B{r}")

        r += 2
        S.subheader(ws, f"A{r}", "Uses")
        r += 1
        first_use = r
        for k, v in su.uses.items():
            S.label_cell(ws, f"A{r}", _humanize(k))
            S.value_cell(ws, f"B{r}", v)
            r += 1
        last_use = r - 1
        S.label_cell(ws, f"A{r}", "Total uses", bold=True)
        S.formula_cell(ws, f"B{r}", f"=SUM(B{first_use}:B{last_use})")
        self._anchors["total_uses"] = _ref(SH_DEAL, f"B{r}")

    # -- Purchase price accounting (live formula walk to goodwill) ------------
    def _ppa(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Purchase price accounting → goodwill (the plug)")
        eqpp = self._anchors["equity_pp"]
        book_eq = self._anchors["assum_tgt_book_eq"]
        exist_gw = self._anchors["assum_tgt_gw"]
        intang = self._anchors["assum_intang_step"]
        ppe = self._anchors["assum_ppe_step"]
        dtl_rate = self._anchors["assum_dtl_rate"]

        S.label_cell(ws, "A3", "Equity purchase price")
        S.formula_cell(ws, "B3", f"={eqpp}")
        S.label_cell(ws, "A4", "Target book equity")
        S.formula_cell(ws, "B4", f"={book_eq}")
        S.label_cell(ws, "A5", "Less: existing goodwill written off")
        S.formula_cell(ws, "B5", f"={exist_gw}")
        S.label_cell(ws, "A6", "Identifiable net assets at book")
        S.formula_cell(ws, "B6", "=B4-B5")
        S.label_cell(ws, "A7", "Intangible step-up")
        S.formula_cell(ws, "B7", f"={intang}")
        S.label_cell(ws, "A8", "PP&E step-up")
        S.formula_cell(ws, "B8", f"={ppe}")
        S.label_cell(ws, "A9", "Deferred tax liability on step-ups")
        S.formula_cell(ws, "B9", f"=(B7+B8)*{dtl_rate}")
        S.label_cell(ws, "A10", "Net identifiable assets", bold=True)
        S.formula_cell(ws, "B10", "=B6+B7+B8-B9")
        S.label_cell(ws, "A11", "Goodwill", bold=True)
        S.formula_cell(ws, "B11", "=B3-B10")
        self._anchors["dtl"] = _ref(SH_PPA, "B9")
        self._anchors["net_ident"] = _ref(SH_PPA, "B10")
        self._anchors["goodwill"] = _ref(SH_PPA, "B11")

    # -- Pro forma income statement (EPS build, primary synergy case) ---------
    def _pro_forma_is(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Pro Forma Income Statement (primary synergy case)")
        combo = model.primary_combination()
        bridge = combo.eps_bridge
        nyr = len(bridge)
        new_sh = self._anchors["new_shares"]

        S.label_cell(ws, "A3", "Line ($)", bold=True)
        for j in range(nyr):
            S.label_cell(ws, f"{_col(2 + j)}3", f"Y{j + 1}", bold=True)

        rows = [
            (4, "Acquirer standalone net income", lambda b: b.acquirer_standalone_ni, "acq_ni"),
            (5, "Target standalone net income", lambda b: b.target_standalone_ni, "tgt_ni"),
            (
                6,
                "Less: incremental interest (after-tax)",
                lambda b: b.incremental_interest_aftertax,
                "int",
            ),
            (
                7,
                "Less: foregone interest income (after-tax)",
                lambda b: b.foregone_interest_aftertax,
                "fore",
            ),
            (8, "Less: incremental D&A (after-tax)", lambda b: b.incremental_da_aftertax, "da"),
            (9, "Plus: synergies (after-tax)", lambda b: b.synergies_aftertax, "syn"),
        ]
        for rr, label, fn, _key in rows:
            S.label_cell(ws, f"A{rr}", label)
            for j in range(nyr):
                S.value_cell(ws, f"{_col(2 + j)}{rr}", fn(bridge[j]))

        S.label_cell(ws, "A10", "Pro forma net income", bold=True)
        S.label_cell(ws, "A11", "Acquirer standalone shares")
        S.label_cell(ws, "A12", "New shares issued")
        S.label_cell(ws, "A13", "Pro forma shares", bold=True)
        S.label_cell(ws, "A14", "Pro forma EPS", bold=True)
        for j in range(nyr):
            c = _col(2 + j)
            # PF net income = acq + tgt - int - foregone - D&A(at) + synergies(at)
            S.formula_cell(ws, f"{c}10", f"={c}4+{c}5-{c}6-{c}7-{c}8+{c}9")
            S.value_cell(ws, f"{c}11", bridge[j].acquirer_standalone_shares, S.FMT_SHARES)
            S.formula_cell(ws, f"{c}12", f"={new_sh}", S.FMT_SHARES)
            S.formula_cell(ws, f"{c}13", f"={c}11+{c}12", S.FMT_SHARES)
            S.formula_cell(ws, f"{c}14", f"={c}10/{c}13", S.FMT_PERSHARE)
            self._anchors[f"pf_ni_{j}"] = f"{c}10"
            self._anchors[f"pf_eps_{j}"] = f"{c}14"
            self._anchors[f"acq_ni_{j}"] = f"{c}4"
            self._anchors[f"acq_sh_{j}"] = f"{c}11"

    def _pro_forma_bs(self, ws: Worksheet, model: MergerModelBundle) -> None:
        """Scaffold: render the primary case's combined balance-sheet lines from
        the engine's projected series (black values). Balancing is an engine
        invariant, verified by the invariant gate — not by this workbook."""
        S.section_header(ws, "A1", "Pro Forma Balance Sheet (engine-computed, invariant-checked)")
        combo = model.primary_combination()
        stmts = combo.proforma_statements
        nyr = len(stmts.periods)
        S.label_cell(ws, "A3", "Line ($)", bold=True)
        for j in range(nyr):
            p = stmts.periods[j]
            fy = p.fy if p.fy is not None else p.end.year
            S.label_cell(ws, f"{_col(2 + j)}3", f"FY{fy}E", bold=True)
        lines = [
            ("Cash & equivalents", LineItem.CASH),
            ("Goodwill", LineItem.GOODWILL),
            ("Total assets", LineItem.TOTAL_ASSETS),
            ("Total liabilities", LineItem.TOTAL_LIABILITIES),
            ("Total stockholders' equity", LineItem.TOTAL_EQUITY),
        ]
        for i, (label, li) in enumerate(lines):
            rr = 4 + i
            S.label_cell(ws, f"A{rr}", label)
            series = stmts.series(li)
            for j in range(nyr):
                v = series[j] if j < len(series) else None
                if v is not None:
                    S.value_cell(ws, f"{_col(2 + j)}{rr}", v)

    # -- Accretion / dilution (EPS bridge → A/D %) ----------------------------
    def _accretion(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Accretion / (Dilution) — primary synergy case")
        combo = model.primary_combination()
        nyr = len(combo.eps_bridge)
        S.label_cell(ws, "A3", "Line", bold=True)
        for j in range(nyr):
            S.label_cell(ws, f"{_col(2 + j)}3", f"Y{j + 1}", bold=True)

        S.label_cell(ws, "A4", "Acquirer standalone EPS")
        S.label_cell(ws, "A5", "Pro forma EPS")
        S.label_cell(ws, "A6", "Accretion / (dilution)", bold=True)
        for j in range(nyr):
            c = _col(2 + j)
            acq_ni = f"'{SH_PF_IS}'!{self._anchors[f'acq_ni_{j}']}"
            acq_sh = f"'{SH_PF_IS}'!{self._anchors[f'acq_sh_{j}']}"
            pf_eps = f"'{SH_PF_IS}'!{self._anchors[f'pf_eps_{j}']}"
            S.formula_cell(ws, f"{c}4", f"={acq_ni}/{acq_sh}", S.FMT_PERSHARE)
            S.formula_cell(ws, f"{c}5", f"={pf_eps}", S.FMT_PERSHARE)
            S.formula_cell(ws, f"{c}6", f"={c}5/{c}4-1", S.FMT_PERCENT)
            self._anchors[f"accr_{j}"] = f"{c}6"

    # -- Contribution & ownership ---------------------------------------------
    def _contribution(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Contribution vs. pro forma ownership")
        combo = model.primary_combination()
        contrib = combo.contribution
        acq_shares = combo.eps_bridge[0].acquirer_standalone_shares
        new_sh = self._anchors["new_shares"]

        S.subheader(ws, "A3", "Contribution ($)")
        rows = [
            (4, "Acquirer revenue", contrib.acquirer_revenue),
            (5, "Target revenue", contrib.target_revenue),
            (6, "Acquirer EBITDA", contrib.acquirer_ebitda),
            (7, "Target EBITDA", contrib.target_ebitda),
            (8, "Acquirer net income", contrib.acquirer_net_income),
            (9, "Target net income", contrib.target_net_income),
        ]
        for rr, label, val in rows:
            S.label_cell(ws, f"A{rr}", label)
            S.value_cell(ws, f"B{rr}", val)

        S.subheader(ws, "A11", "Pro forma ownership")
        S.label_cell(ws, "A12", "Acquirer existing shares")
        S.value_cell(ws, "B12", acq_shares, S.FMT_SHARES)
        S.label_cell(ws, "A13", "New shares issued to target")
        S.formula_cell(ws, "B13", f"={new_sh}", S.FMT_SHARES)
        S.label_cell(ws, "A14", "Acquirer ownership", bold=True)
        S.formula_cell(ws, "B14", "=B12/(B12+B13)", S.FMT_PERCENT)
        S.label_cell(ws, "A15", "Target ownership", bold=True)
        S.formula_cell(ws, "B15", "=B13/(B12+B13)", S.FMT_PERCENT)
        self._anchors["acq_own"] = _ref(SH_CONTRIB, "B14")
        self._anchors["tgt_own"] = _ref(SH_CONTRIB, "B15")

    # -- Sensitivities (premium × synergies grid) -----------------------------
    def _sensitivities(self, ws: Worksheet, model: MergerModelBundle) -> None:
        """Premium × synergies accretion/(dilution) grid.

        Which cells are LIVE (verified by the differential) vs. engine-value:

        * **Live formulas** — every interior grid cell in a synergy column with
          index ≥ 2 (i.e. columns D onward). Because Year-N A/D is *exactly*
          linear in the synergy run-rate (the combination engine adds
          after-tax synergies additively), each such cell is a live linear
          interpolation off the row's two anchor columns (B, C) and the synergy
          axis header row — a pure function of cells already on the sheet, so it
          recomputes to the engine grid value to the cent.
        * **Engine values** — the two anchor columns (B, C) of each row, the row
          (premium) and column (synergy) axis labels, and the breakeven cell.
          The breakeven synergies figure is the engine's bisection root; it is
          NOT a clean function of the grid cells (reconstructing it by
          interpolation drifts past the cent tolerance), so it is written as a
          labeled engine value and is not part of the differential.
        """
        S.section_header(ws, "A1", "Sensitivities — premium × synergies (Year-1 A/D)")
        sens = model.sensitivities
        # Always define a breakeven-synergies cell so the named range resolves.
        S.label_cell(ws, "A3", "Breakeven synergies (Year-1 A/D = 0)", bold=True)
        breakeven = sens.breakeven_synergies if sens is not None else None
        S.value_cell(ws, "B3", breakeven if breakeven is not None else 0.0)
        self._anchors["breakeven"] = _ref(SH_SENS, "B3")
        if sens is None:
            S.label_cell(ws, "A5", "Grid populated at G2 once the scenario engine runs.")
            return
        grid = sens.premium_x_synergies
        top = 6
        cv = grid.col_values
        # A live linear interpolation needs two distinct synergy anchors.
        interp_ok = len(cv) >= 2 and (cv[1] - cv[0]) != 0
        syn0 = f"{_col(2)}{top}"  # first synergy axis header (anchor 0)
        syn1 = f"{_col(3)}{top}"  # second synergy axis header (anchor 1)
        S.label_cell(ws, f"A{top}", f"{grid.row_label} \\ {grid.col_label}", bold=True)
        for cidx, cval in enumerate(cv):
            S.value_cell(ws, f"{_col(2 + cidx)}{top}", cval, S.FMT_CURRENCY)
        for ridx, rv in enumerate(grid.row_values):
            rr = top + 1 + ridx
            S.value_cell(ws, f"A{rr}", rv, S.FMT_PERCENT)
            row = grid.values[ridx]
            a0 = f"{_col(2)}{rr}"  # anchor cell at synergy col 0
            a1 = f"{_col(3)}{rr}"  # anchor cell at synergy col 1
            can_interp = interp_ok and row[0] is not None and row[1] is not None
            for cidx in range(len(cv)):
                gval = row[cidx] if cidx < len(row) else None
                coord = f"{_col(2 + cidx)}{rr}"
                if cidx >= 2 and can_interp and gval is not None:
                    synj = f"{_col(2 + cidx)}{top}"
                    S.formula_cell(
                        ws,
                        coord,
                        f"={a0}+({a1}-{a0})/({syn1}-{syn0})*({synj}-{syn0})",
                        S.FMT_PERCENT,
                    )
                elif gval is not None:
                    S.value_cell(ws, coord, gval, S.FMT_PERCENT)

    def _precedents(self, ws: Worksheet, model: MergerModelBundle) -> None:
        S.section_header(ws, "A1", "Precedent Transactions (premiums / multiples — cited)")
        headers = ["Date", "Acquirer", "Target", "EV", "EV/Rev", "EV/EBITDA", "Source"]
        for k, h in enumerate(headers):
            S.subheader(ws, f"{_col(1 + k)}3", h)
        r = 4
        for t in model.precedents:
            S.label_cell(ws, f"A{r}", t.date)
            S.label_cell(ws, f"B{r}", t.acquirer)
            S.label_cell(ws, f"C{r}", t.target)
            S.value_cell(ws, f"D{r}", t.ev)
            _mult_or_blank(ws, f"E{r}", t.ev_revenue)
            _mult_or_blank(ws, f"F{r}", t.ev_ebitda)
            S.label_cell(ws, f"G{r}", t.source)
            r += 1
        if not model.precedents:
            S.label_cell(ws, "A4", "Curated precedent set added at G2/G3.")

    def _fairness(self, ws: Worksheet, model: MergerModelBundle) -> None:
        """Advisor disclosed implied ranges vs. our reproduction.

        Which cells are LIVE (verified by the differential) vs. engine-value:

        * **Engine/extraction values** — the four range endpoints per method
          (disclosed low/high from the proxy, our low/high from our engine). A
          range endpoint is an output, correctly written as a value.
        * **Live formulas** — three per method, each a pure function of the
          endpoint cells on the same row: the disclosed midpoint ``(low+high)/2``
          (col G), our midpoint (col H), and the overlap %
          ``|intersection| / |disclosed width|`` clamped to [0, 1] (col I). These
          recompute to the engine's midpoints and ``overlap_pct`` to the cent.
          A formula is emitted only when its inputs are all present on the row.
        """
        S.section_header(ws, "A1", "Fairness Comparison — advisor ranges vs. our reproduction")
        headers = [
            "Advisor",
            "Method",
            "Disclosed low",
            "Disclosed high",
            "Our low",
            "Our high",
            "Disclosed mid",
            "Our mid",
            "Overlap %",
        ]
        for k, h in enumerate(headers):
            S.subheader(ws, f"{_col(1 + k)}3", h)
        fd = model.fairness_differential
        if fd is None or not fd.reproductions:
            S.label_cell(ws, "A4", "Fairness differential populated at W2/G2 from the proxy.")
            return
        r = 4
        for rep in fd.reproductions:
            S.label_cell(ws, f"A{r}", rep.advisor)
            S.label_cell(ws, f"B{r}", rep.method)
            dlo, dhi, olo, ohi = (
                f"C{r}",
                f"D{r}",
                f"E{r}",
                f"F{r}",
            )
            for coord, v in (
                (dlo, rep.disclosed_low),
                (dhi, rep.disclosed_high),
                (olo, rep.our_low),
                (ohi, rep.our_high),
            ):
                if v is not None:
                    S.value_cell(ws, coord, v, S.FMT_PERSHARE)
            # Disclosed midpoint — live off the disclosed low/high cells.
            if rep.disclosed_low is not None and rep.disclosed_high is not None:
                S.formula_cell(ws, f"G{r}", f"=({dlo}+{dhi})/2", S.FMT_PERSHARE)
            # Our midpoint — live off our low/high cells.
            if rep.our_low is not None and rep.our_high is not None:
                S.formula_cell(ws, f"H{r}", f"=({olo}+{ohi})/2", S.FMT_PERSHARE)
            # Overlap % — |intersection| / |disclosed width|, clamped to [0, 1];
            # mirrors src.fairness.engine._overlap_pct exactly. A non-None
            # overlap_pct guarantees all four endpoints are present and the
            # disclosed width is positive (so the formula's denominator is safe).
            if rep.overlap_pct is not None:
                S.formula_cell(
                    ws,
                    f"I{r}",
                    f"=MAX(0,MIN(1,(MIN({dhi},{ohi})-MAX({dlo},{olo}))/({dhi}-{dlo})))",
                    S.FMT_PERCENT,
                )
            r += 1

    # -- Cover (headline links to the model) ----------------------------------
    def _cover(self, ws: Worksheet, model: MergerModelBundle) -> None:
        t = model.terms
        S.section_header(
            ws,
            "A1",
            f"{t.acquirer_name} ({t.acquirer_ticker}) acquires {t.target_name} ({t.target_ticker})",
        )
        S.label_cell(ws, "A3", "Announced")
        ws["B3"].value = t.announce_date.isoformat()
        S.label_cell(ws, "A4", "Expected close")
        ws["B4"].value = t.close_date.isoformat() if t.close_date else "Pending"

        S.subheader(ws, "A6", "Headline")
        S.label_cell(ws, "A7", "Equity purchase price")
        S.formula_cell(ws, "B7", f"={self._anchors['equity_pp']}")
        S.label_cell(ws, "A8", "Implied premium")
        S.formula_cell(ws, "B8", f"={self._anchors['implied_premium']}", S.FMT_PERCENT)
        S.label_cell(ws, "A9", "Goodwill created")
        S.formula_cell(ws, "B9", f"={self._anchors['goodwill']}")

        S.subheader(ws, "A11", "Accretion / (dilution) — primary case")
        nyr = len(model.primary_combination().eps_bridge)
        for j in range(min(nyr, 3)):
            rr = 12 + j
            S.label_cell(ws, f"A{rr}", f"Year {j + 1}")
            S.formula_cell(ws, f"B{rr}", f"='{SH_AD}'!{self._anchors[f'accr_{j}']}", S.FMT_PERCENT)

        S.label_cell(
            ws,
            "A16",
            "Educational project — not investment advice. SEC EDGAR data used per "
            "fair-access policy; no SEC or advisor endorsement implied.",
        )

    # -- Native charts + ergonomics -------------------------------------------
    def _add_charts(self, sheets: dict[str, Worksheet], model: MergerModelBundle) -> None:
        """One banker-grade chart: accretion/(dilution) by year (a live bar off
        the Accretion-Dilution tab). Charts do not affect recalculation."""
        nyr = len(model.primary_combination().eps_bridge)
        if nyr < 1:
            return
        ad = sheets[SH_AD]
        bar = BarChart()
        bar.type = "col"
        bar.title = "Accretion / (Dilution) by Year"
        bar.height, bar.width = 6.5, 15
        data = Reference(ad, min_col=2, max_col=1 + nyr, min_row=6, max_row=6)
        cats = Reference(ad, min_col=2, max_col=1 + nyr, min_row=3, max_row=3)
        bar.add_data(data, from_rows=True, titles_from_data=False)
        bar.set_categories(cats)
        ad.add_chart(bar, "A9")

    def _apply_ergonomics(self, sheets: dict[str, Worksheet]) -> None:
        for name, ws in sheets.items():
            ws.column_dimensions["A"].width = _LABEL_COL_WIDTH
            for col in ("B", "C", "D", "E", "F", "G", "H", "I", "J"):
                ws.column_dimensions[col].width = _DATA_COL_WIDTH
            ws.freeze_panes = "B4" if name in _GRID_TABS else "B1"

    def _define_names(self, wb: Workbook) -> None:
        a = self._anchors
        # Accretion year cells clamp to the last available year when fewer than 3
        # projected years exist, so every named range resolves to one cell.
        accr_years = sorted(int(k.split("_")[1]) for k in a if k.startswith("accr_"))
        last_accr = accr_years[-1] if accr_years else 0

        def accr_ref(j: int) -> str:
            idx = j if j in accr_years else last_accr
            return f"'{SH_AD}'!{a[f'accr_{idx}']}"

        mapping = {
            "DealValue": a["total_uses"],
            "EquityPurchasePrice": a["equity_pp"],
            "CashPerShare": a["assum_cash_ps"],
            "ExchangeRatio": a["assum_xratio"],
            "ImpliedPremium": a["implied_premium"],
            "Goodwill": a["goodwill"],
            "NewDebt": a["assum_new_debt"],
            "NewSharesIssued": a["new_shares"],
            "ProFormaEPS_Y1": f"'{SH_PF_IS}'!{a['pf_eps_0']}",
            "Accretion_Y1": accr_ref(0),
            "Accretion_Y2": accr_ref(1),
            "Accretion_Y3": accr_ref(2),
            "BreakevenSynergies": a["breakeven"],
            "AcquirerOwnership": a["acq_own"],
            "TargetOwnership": a["tgt_own"],
        }
        for name, ref in mapping.items():
            wb.defined_names[name] = DefinedName(name, attr_text=ref)

    @staticmethod
    def _save_pinned(wb: Workbook, path: str) -> None:
        """Save so a rebuild is byte-identical across processes and time.

        Two openpyxl behaviors otherwise defeat determinism: (1) ``Workbook.save``
        re-stamps ``properties.modified`` with ``datetime.now()`` at write time,
        and (2) every ZIP entry gets the current wall-clock time. Driving
        :class:`ExcelWriter` directly avoids (1); a ZipFile that pins each
        entry's ``date_time`` to a fixed epoch avoids (2). Ported from thesis."""
        from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

        from openpyxl.writer.excel import ExcelWriter

        fixed_date_time = (2026, 1, 1, 0, 0, 0)

        class _PinnedZip(ZipFile):
            def writestr(self, zinfo_or_arcname, data, *args, **kwargs):
                if isinstance(zinfo_or_arcname, str):
                    zinfo_or_arcname = ZipInfo(zinfo_or_arcname, date_time=fixed_date_time)
                    zinfo_or_arcname.compress_type = self.compression
                else:
                    zinfo_or_arcname.date_time = fixed_date_time
                return super().writestr(zinfo_or_arcname, data, *args, **kwargs)

            def write(self, filename, arcname=None, *args, **kwargs):
                with open(filename, "rb") as fh:
                    data = fh.read()
                return self.writestr(arcname if arcname is not None else filename, data)

        archive = _PinnedZip(path, "w", ZIP_DEFLATED, allowZip64=True)
        try:
            ExcelWriter(wb, archive).save()
        finally:
            archive.close()


def build_verifier_cell_map(model: MergerModelBundle) -> dict[tuple[str, str], float]:
    """Authoritative ``(sheet, coord) -> engine-value`` map for the differential.

    The verifier recalculates the workbook (via the ``formulas`` library) and
    asserts each of these cells equals the mapped engine value to the cent. Every
    listed cell is a LIVE FORMULA that reproduces an engine output: the
    consideration build + S&U totals, the PPA walk to goodwill, the standalone
    revenue projections, the per-year pro forma net income / EPS, the
    accretion/(dilution) %, the ownership split, the Sensitivities grid interior
    (synergy-column index ≥ 2, a live linear interpolation off the row's two
    engine-value anchors — exact because A/D is linear in synergies), and the
    Fairness Comparison disclosed/our midpoints + per-method overlap %.
    """
    m: dict[tuple[str, str], float] = {}
    cons = model.deal.consideration
    ppa = model.deal.ppa
    su = model.deal.sources_and_uses

    # Deal & S&U (consideration build + balancing totals).
    m[(SH_DEAL, "B4")] = cons.aggregate_cash
    m[(SH_DEAL, "B5")] = cons.aggregate_stock_value
    m[(SH_DEAL, "B6")] = cons.equity_purchase_price
    m[(SH_DEAL, "B8")] = cons.new_shares_issued
    if cons.implied_premium_pct is not None:
        m[(SH_DEAL, "B9")] = cons.implied_premium_pct
    m[(SH_DEAL, _coord_of(model, "total_sources"))] = su.total_sources
    m[(SH_DEAL, _coord_of(model, "total_uses"))] = su.total_uses

    # PPA walk.
    m[(SH_PPA, "B9")] = ppa.deferred_tax_liability
    m[(SH_PPA, "B10")] = ppa.net_identifiable_assets
    m[(SH_PPA, "B11")] = ppa.goodwill

    # Standalone revenue projections (live chain off Assumptions growth inputs).
    for prefix, stmts in (("acq", model.acquirer_statements), ("tgt", model.target_statements)):
        n = stmts.n_hist
        proj = stmts.series(LineItem.REVENUE)[n:]
        sheet = SH_ACQ if prefix == "acq" else SH_TGT
        for j, v in enumerate(proj):
            if v is not None:
                m[(sheet, f"{_col(3 + j)}4")] = v

    # Pro forma IS + accretion, per projected year (primary synergy case).
    combo = model.primary_combination()
    for j, b in enumerate(combo.eps_bridge):
        c = _col(2 + j)
        m[(SH_PF_IS, f"{c}10")] = b.pro_forma_net_income
        m[(SH_PF_IS, f"{c}14")] = b.pro_forma_eps
        accr = b.accretion_dilution_pct
        if accr is not None:
            m[(SH_AD, f"{c}6")] = accr

    # Ownership split.
    m[(SH_CONTRIB, "B14")] = combo.contribution.acquirer_ownership_pct
    m[(SH_CONTRIB, "B15")] = combo.contribution.target_ownership_pct

    # Sensitivities grid interior (synergy-column index ≥ 2): a live linear
    # interpolation off each row's two anchor columns. A/D is exactly linear in
    # the synergy run-rate, so the formula reproduces the engine grid to the cent.
    sens = model.sensitivities
    if sens is not None:
        grid = sens.premium_x_synergies
        top = 6
        cv = grid.col_values
        interp_ok = len(cv) >= 2 and (cv[1] - cv[0]) != 0
        for ridx, row in enumerate(grid.values):
            rr = top + 1 + ridx
            if not (interp_ok and row[0] is not None and row[1] is not None):
                continue
            for cidx in range(2, len(cv)):
                gval = row[cidx] if cidx < len(row) else None
                if gval is not None:
                    m[(SH_SENS, f"{_col(2 + cidx)}{rr}")] = gval

    # Fairness Comparison: disclosed/our midpoints (cols G/H) + overlap % (col I),
    # each a live formula off the range-endpoint cells on the same row.
    fd = model.fairness_differential
    if fd is not None:
        for i, rep in enumerate(fd.reproductions):
            rr = 4 + i
            if rep.disclosed_low is not None and rep.disclosed_high is not None:
                m[(SH_FAIR, f"G{rr}")] = (rep.disclosed_low + rep.disclosed_high) / 2.0
            if rep.our_low is not None and rep.our_high is not None:
                m[(SH_FAIR, f"H{rr}")] = (rep.our_low + rep.our_high) / 2.0
            if rep.overlap_pct is not None:
                m[(SH_FAIR, f"I{rr}")] = rep.overlap_pct
    return m


# ---------------------------------------------------------------------------
# Small helpers (ported / adapted from the thesis writer).
# ---------------------------------------------------------------------------
def _coord_of(model: MergerModelBundle, key: str) -> str:
    """Row of a Sources/Uses total, computed the same way the writer lays it out.

    The S&U totals sit one row below the last line of each block; deriving the
    coordinate here (rather than threading writer state into the map) keeps
    build_verifier_cell_map a pure function of the bundle."""
    su = model.deal.sources_and_uses
    # Sources occupy rows 12..(11 + n_src); the total sits at 12 + n_src.
    total_sources_row = 12 + len(su.sources)
    if key == "total_sources":
        return f"B{total_sources_row}"
    # After the sources total: +2 gap rows, +1 subheader → first use row.
    first_use = total_sources_row + 3
    return f"B{first_use + len(su.uses)}"


def _mult_or_blank(ws: Worksheet, coord: str, v: float | None) -> None:
    if v is None:
        return
    S.value_cell(ws, coord, v, S.FMT_MULTIPLE)


def _humanize(key: str) -> str:
    return key.replace("_", " ").strip().capitalize()


def _sv(sourced) -> float | None:
    """Value of a SourcedValue, or None (honest unknown / absent)."""
    return None if sourced is None else sourced.value


def _first_not_none(*vals):
    for v in vals:
        if v is not None:
            return v
    return None


def _safe_div(a: float | None, b: float | None) -> float | None:
    if a is None or b in (None, 0.0):
        return None
    return a / b


def _col(idx: int) -> str:
    """1-based column index -> Excel column letter(s)."""
    letters = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _last_hist(stmts, li: LineItem) -> float | None:
    series = stmts.series(li)[: stmts.n_hist]
    for v in reversed(series):
        if v is not None:
            return v
    return None


def _ref(sheet: str, coord: str) -> str:
    """A cross-sheet qualified reference, e.g. ``'Deal & S&U'!B6``."""
    return f"'{sheet}'!{coord}"
