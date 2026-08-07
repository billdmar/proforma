"""Tests for the live-formula merger-model workbook writer.

The headline property (the moat): every cell in ``build_verifier_cell_map`` is a
live Excel FORMULA that, when the file is recalculated with the ``formulas``
library, reproduces the engine value in the ``MergerModelBundle`` to the cent.
This test file owns its inputs: ``_sample_bundle()`` hand-builds a small but
complete bundle so the writer can be exercised without importing any sibling
engine module.
"""

from __future__ import annotations

import os
from datetime import date

import formulas
import pytest
from openpyxl import load_workbook
from src.interfaces import (
    CombinationResult,
    Consideration,
    ContributionAnalysis,
    DealAssumptions,
    DealResult,
    EPSBridge,
    FairnessDifferentialReport,
    MergerModelBundle,
    MethodologyReproduction,
    PrecedentTransaction,
    ProjectionAssumptions,
    PurchasePriceAllocation,
    SensitivityGrid,
    SensitivitySet,
    SourcesAndUses,
    StatementSet,
    SynergyCase,
)
from src.schema import (
    CompanyMeta,
    ConsiderationType,
    DealTerms,
    DocProvenance,
    LineItem,
    Period,
    PeriodType,
    SourcedValue,
)
from src.workbook import ExcelWorkbookWriter, build_verifier_cell_map
from src.workbook.styles import formula_cell
from src.workbook.writer import NAMED_RANGES, SHEET_ORDER


# ---------------------------------------------------------------------------
# Hand-built, self-consistent bundle (own test inputs; no engine imports).
# ---------------------------------------------------------------------------
def _dur(y: int) -> Period:
    return Period(PeriodType.DURATION, end=date(y, 12, 31), start=date(y, 1, 1), fy=y, fp="FY")


def _statements(base_rev: float, growth: list[float], ni: list[float]) -> StatementSet:
    """One historical year + len(growth) projected years, revenue chained by the
    same growth the workbook applies, so the workbook's live chain reproduces the
    engine's projected revenue exactly."""
    periods = [_dur(2024)] + [_dur(2025 + j) for j in range(len(growth))]
    rev = [base_rev]
    for g in growth:
        rev.append(rev[-1] * (1 + g))
    rows: dict[LineItem, list[float | None]] = {
        LineItem.REVENUE: rev,
        LineItem.NET_INCOME: [None] + ni,
        LineItem.CASH: [None] + [1000.0] * len(growth),
        LineItem.GOODWILL: [None] + [5000.0] * len(growth),
        LineItem.TOTAL_ASSETS: [None] + [20000.0] * len(growth),
        LineItem.TOTAL_LIABILITIES: [None] + [8000.0] * len(growth),
        LineItem.TOTAL_EQUITY: [None] + [12000.0] * len(growth),
    }
    return StatementSet(periods=periods, rows=rows, n_hist=1)


def _sample_bundle() -> MergerModelBundle:
    acquirer = CompanyMeta(cik="0000883241", ticker="SNPS", name="Synopsys, Inc.")
    target = CompanyMeta(cik="0001013462", ticker="ANSS", name="ANSYS, Inc.")

    n = 3  # projected years
    acq_growth = [0.10, 0.09, 0.08]
    tgt_growth = [0.08, 0.07, 0.06]
    acq_stmts = _statements(6000.0, acq_growth, ni=[1200.0, 1320.0, 1450.0])
    tgt_stmts = _statements(2200.0, tgt_growth, ni=[500.0, 540.0, 580.0])
    acq_assum = ProjectionAssumptions(n_years=n, revenue_growth=acq_growth)
    tgt_assum = ProjectionAssumptions(n_years=n, revenue_growth=tgt_growth)

    # --- Deal terms & our assumptions -------------------------------------
    cash_ps = 197.0
    xratio = 0.3450
    ref_px = 500.0
    tgt_shares = 87_000_000.0
    prem_ref = 315.00
    terms = DealTerms(
        acquirer_ticker="SNPS",
        target_ticker="ANSS",
        acquirer_name="Synopsys, Inc.",
        target_name="ANSYS, Inc.",
        announce_date=date(2024, 1, 16),
        close_date=None,
        consideration_type=ConsiderationType.MIXED,
        premium_reference_price=SourcedValue(
            value=prem_ref,
            provenance=DocProvenance(
                accession="0000000000-24-000000",
                form="DEFM14A",
                filed=date(2024, 3, 1),
                section="Background of the Merger — unaffected price",
            ),
            unit="USD/share",
            label="unaffected price",
        ),
    )

    # Consideration build (matches the workbook's formulas exactly).
    agg_cash = cash_ps * tgt_shares
    agg_stock = xratio * ref_px * tgt_shares
    eqpp = agg_cash + agg_stock
    total_ps = cash_ps + xratio * ref_px
    new_shares = xratio * tgt_shares
    premium = total_ps / prem_ref - 1.0
    cons = Consideration(
        cash_per_share=cash_ps,
        stock_per_share_value=xratio * ref_px,
        total_per_share=total_ps,
        target_shares=tgt_shares,
        aggregate_cash=agg_cash,
        aggregate_stock_value=agg_stock,
        equity_purchase_price=eqpp,
        implied_premium_pct=premium,
        exchange_ratio=xratio,
        new_shares_issued=new_shares,
    )

    da = DealAssumptions(
        new_debt=16_000_000_000.0,
        new_debt_rate=0.05,
        cash_on_hand_used=3_000_000_000.0,
        foregone_cash_yield=0.03,
        advisory_fees=200_000_000.0,
        financing_fees=150_000_000.0,
        intangible_step_up=8_000_000_000.0,
        ppe_step_up=500_000_000.0,
        deferred_tax_rate=0.21,
        marginal_tax_rate=0.21,
        new_shares_issued=new_shares,
    )

    # Sources & uses (must balance; totals verified as live SUMs).
    stock_issued = agg_stock
    fees = da.advisory_fees + da.financing_fees
    total = agg_cash + stock_issued + fees
    su = SourcesAndUses(
        sources={
            "new_debt": da.new_debt,
            "cash_on_hand": da.cash_on_hand_used,
            "stock_issued": stock_issued,
            "revolver_plug": total - da.new_debt - da.cash_on_hand_used - stock_issued,
        },
        uses={
            "equity_purchase_price": eqpp,
            "transaction_fees": fees,
        },
    )
    assert su.balances()

    # PPA walk.
    book_eq = 6_000_000_000.0
    exist_gw = 3_500_000_000.0
    ident_book = book_eq - exist_gw
    dtl = (da.intangible_step_up + da.ppe_step_up) * da.deferred_tax_rate
    net_ident = ident_book + da.intangible_step_up + da.ppe_step_up - dtl
    goodwill = eqpp - net_ident
    ppa = PurchasePriceAllocation(
        equity_purchase_price=eqpp,
        target_book_equity=book_eq,
        target_existing_goodwill=exist_gw,
        identifiable_net_assets_at_book=ident_book,
        intangible_step_up=da.intangible_step_up,
        ppe_step_up=da.ppe_step_up,
        deferred_tax_liability=dtl,
        net_identifiable_assets=net_ident,
        goodwill=goodwill,
    )
    assert ppa.ties()

    deal = DealResult(consideration=cons, sources_and_uses=su, ppa=ppa, incremental_da_annual=0.0)

    # --- Combination: primary (management) case + a conservative case -----
    acq_ni = [1200.0e6, 1320.0e6, 1450.0e6]
    tgt_ni = [500.0e6, 540.0e6, 580.0e6]
    acq_shares = 152_000_000.0
    pf_shares = acq_shares + new_shares
    tax = da.marginal_tax_rate

    def _bridges(run_rate: float, phase: list[float]) -> list[EPSBridge]:
        out = []
        for j in range(n):
            inc_int = da.new_debt * da.new_debt_rate
            fore = da.cash_on_hand_used * da.foregone_cash_yield
            inc_da_at = 0.0
            syn_at = run_rate * phase[j] * (1 - tax)
            pf_ni = acq_ni[j] + tgt_ni[j] - inc_int - fore - inc_da_at + syn_at
            acq_eps = acq_ni[j] / acq_shares
            pf_eps = pf_ni / pf_shares
            out.append(
                EPSBridge(
                    year=_dur(2025 + j),
                    acquirer_standalone_ni=acq_ni[j],
                    target_standalone_ni=tgt_ni[j],
                    incremental_interest_expense=inc_int,
                    foregone_interest_income=fore,
                    incremental_da_aftertax=inc_da_at,
                    synergies_aftertax=syn_at,
                    pro_forma_net_income=pf_ni,
                    acquirer_standalone_shares=acq_shares,
                    new_shares_issued=new_shares,
                    pro_forma_shares=pf_shares,
                    acquirer_standalone_eps=acq_eps,
                    pro_forma_eps=pf_eps,
                )
            )
        return out

    contrib = ContributionAnalysis(
        acquirer_revenue=6000.0e6,
        target_revenue=2200.0e6,
        acquirer_ebitda=2000.0e6,
        target_ebitda=800.0e6,
        acquirer_net_income=1200.0e6,
        target_net_income=500.0e6,
        acquirer_ownership_pct=acq_shares / pf_shares,
        target_ownership_pct=new_shares / pf_shares,
    )
    pf_bs = _statements(8200.0e6, [0.09, 0.08, 0.07], ni=[1.0, 1.0, 1.0])

    mgmt = CombinationResult(
        synergy_case=SynergyCase(
            name="Management",
            run_rate_annual=400.0e6,
            phase_in=[0.33, 0.66, 1.0],
            is_disclosed=True,
        ),
        proforma_statements=pf_bs,
        eps_bridge=_bridges(400.0e6, [0.33, 0.66, 1.0]),
        contribution=contrib,
    )
    ours = CombinationResult(
        synergy_case=SynergyCase(
            name="Conservative",
            run_rate_annual=250.0e6,
            phase_in=[0.25, 0.50, 0.90],
            is_disclosed=False,
        ),
        proforma_statements=pf_bs,
        eps_bridge=_bridges(250.0e6, [0.25, 0.50, 0.90]),
        contribution=contrib,
    )

    return MergerModelBundle(
        acquirer=acquirer,
        target=target,
        terms=terms,
        acquirer_statements=acq_stmts,
        target_statements=tgt_stmts,
        acquirer_assumptions=acq_assum,
        target_assumptions=tgt_assum,
        deal_assumptions=da,
        deal=deal,
        combinations=[mgmt, ours],
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------
def test_write_roundtrip_sheets_and_order(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _sample_bundle())
    wb = load_workbook(p)
    assert wb.sheetnames == SHEET_ORDER
    assert len(wb.sheetnames) == 13


def test_named_ranges_resolve_to_single_cell(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _sample_bundle())
    wb = load_workbook(p)
    for name in NAMED_RANGES:
        assert name in wb.defined_names, f"missing named range {name}"
        dests = list(wb.defined_names[name].destinations)
        assert len(dests) == 1, f"{name} must resolve to one cell"


def test_formula_cell_rejects_non_formula():
    from openpyxl import Workbook

    ws = Workbook().active
    with pytest.raises(ValueError):
        formula_cell(ws, "A1", "1234")  # not a formula → must raise


def test_verifier_cells_are_formulas(tmp_path):
    p = tmp_path / "m.xlsx"
    bundle = _sample_bundle()
    ExcelWorkbookWriter().write(str(p), bundle)
    wb = load_workbook(p, data_only=False)
    cell_map = build_verifier_cell_map(bundle)
    assert len(cell_map) >= 8
    for sheet, coord in cell_map:
        c = wb[sheet][coord]
        assert c.data_type == "f", f"{sheet}!{coord} must be a formula, got {c.value!r}"


def test_the_differential_recalc_to_the_cent(tmp_path):
    """THE moat: recalc the workbook with the ``formulas`` library and assert
    every mapped cell reproduces the engine value to the cent (tol=0.01)."""
    p = tmp_path / "m.xlsx"
    bundle = _sample_bundle()
    ExcelWorkbookWriter().write(str(p), bundle)
    cell_map = build_verifier_cell_map(bundle)
    assert len(cell_map) >= 8

    sol = formulas.ExcelModel().loads(str(p)).finish().calculate()
    fname = os.path.basename(str(p))
    mismatches = []
    for (sheet, coord), engine_value in cell_map.items():
        node = sol.get(f"'[{fname}]{sheet.upper()}'!{coord}")
        wb_value = None if node is None else float(node.value[0, 0])
        if wb_value is None or abs(wb_value - engine_value) > 0.01:
            mismatches.append((sheet, coord, engine_value, wb_value))
    assert not mismatches, f"differential mismatches: {mismatches}"


def _enrich(bundle: MergerModelBundle) -> MergerModelBundle:
    """Populate the scaffolded tabs (sensitivities / precedents / fairness) so
    their rendering paths are exercised."""
    bundle.sensitivities = SensitivitySet(
        premium_x_synergies=SensitivityGrid(
            year_idx=0,
            row_label="premium",
            col_label="synergies",
            row_values=[0.20, 0.30, 0.40],
            col_values=[200.0e6, 400.0e6, 600.0e6],
            values=[[0.05, 0.08, 0.11], [0.02, 0.05, 0.08], [None, 0.02, 0.05]],
        ),
        consideration_mix=SensitivityGrid(
            year_idx=0,
            row_label="cash%",
            col_label="synergies",
            row_values=[0.5],
            col_values=[400.0e6],
            values=[[0.05]],
        ),
        breakeven_synergies=350.0e6,
        breakeven_year_idx=0,
    )
    bundle.precedents = [
        PrecedentTransaction(
            date="2023",
            acquirer="Broadcom",
            target="VMware",
            ev=69.0e9,
            ev_revenue=5.2,
            ev_ebitda=None,
            source="press release 2023",
        ),
    ]
    bundle.fairness_differential = FairnessDifferentialReport(
        reproductions=[
            MethodologyReproduction(
                advisor="Evercore",
                method="DCF",
                disclosed_low=300.0,
                disclosed_high=360.0,
                our_low=305.0,
                our_high=None,
                overlap_pct=0.85,
            ),
        ],
    )
    return bundle


def test_scaffold_tabs_render_when_populated(tmp_path):
    p = tmp_path / "m.xlsx"
    ExcelWorkbookWriter().write(str(p), _enrich(_sample_bundle()))
    wb = load_workbook(p)
    assert "BreakevenSynergies" in wb.defined_names
    # Sensitivities grid centre cell populated; precedents + fairness rows filled.
    assert wb["Sensitivities"]["C7"].value is not None
    assert wb["Precedents"]["A4"].value == "2023"
    assert wb["Fairness Comparison"]["A4"].value == "Evercore"


def test_deterministic_byte_identical(tmp_path):
    b = _sample_bundle()
    p1 = tmp_path / "a.xlsx"
    p2 = tmp_path / "b.xlsx"
    ExcelWorkbookWriter().write(str(p1), b)
    ExcelWorkbookWriter().write(str(p2), b)
    assert p1.read_bytes() == p2.read_bytes()
