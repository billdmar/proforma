"""Tests for the Excel banker-convention audit (src/verify/audit.py).

Runs the audit against the real flagship merger workbook (built from committed
fixtures) with and without the model, and against deliberately-corrupted copies
that inject a blue input off the Assumptions tab, bake a value into a formula
region, drop a required named range, and repoint a name at a multi-cell range.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from src.flagship import build_flagship_bundle
from src.verify.audit import AuditReport, audit_workbook
from src.workbook import ExcelWorkbookWriter
from src.workbook.styles import BLUE

_PRECEDENTS_CSV = "data/curated/precedents_software.csv"


def _write(tmp_path) -> str:
    bundle = build_flagship_bundle(_PRECEDENTS_CSV)
    path = tmp_path / "SNPS_ANSS_model.xlsx"
    ExcelWorkbookWriter().write(str(path), bundle)
    return str(path)


def test_audit_passes_with_model_on_compliant_workbook(tmp_path):
    # The writer emits blue inputs ONLY on the Assumptions tab and every mapped
    # cell as a live formula, so its unmodified output passes the full audit.
    path = _write(tmp_path)
    report = audit_workbook(path, model=build_flagship_bundle(_PRECEDENTS_CSV))
    assert report.passed, report.violations
    # Formula-region check (mapped cells) + blue-font + 15 named ranges all ran.
    assert report.cells_checked > 15


def test_audit_passes_without_model_on_compliant_workbook(tmp_path):
    path = _write(tmp_path)
    report = audit_workbook(path)  # skips the formula-region sub-check
    assert report.passed, report.violations


def test_audit_flags_baked_value_in_formula_region(tmp_path):
    # Overwrite a mapped formula cell (the goodwill plug) with a baked number.
    path = _write(tmp_path)
    wb = load_workbook(path)
    wb["PPA"]["B11"] = 42.0
    wb.save(path)

    report = audit_workbook(path, model=build_flagship_bundle(_PRECEDENTS_CSV))
    assert not report.passed
    assert any("PPA!B11" in v and "baked value" in v for v in report.violations)


def test_audit_flags_blue_input_off_assumptions(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    # Inject a blue (input) number onto the PPA tab — a banker-convention breach.
    cell = wb["PPA"]["E30"]
    cell.value = 123.0
    cell.font = Font(color=BLUE)
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("blue input" in v and "PPA!E30" in v for v in report.violations)


def test_audit_flags_missing_named_range(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    del wb.defined_names["Goodwill"]
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("missing named range 'Goodwill'" in v for v in report.violations)


def test_audit_flags_named_range_spanning_a_range(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    # Repoint a required name at a multi-cell range — the audit wants one cell.
    wb.defined_names["Goodwill"] = DefinedName("Goodwill", attr_text="'PPA'!$B$11:$B$12")
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("spans a range" in v for v in report.violations)


def test_audit_allows_blue_input_on_cover_metadata_cell(tmp_path):
    # A blue input on an allowed Cover metadata cell (B3) is exempt — the audit
    # still passes (this locks the Cover-exemption branch).
    path = _write(tmp_path)
    wb = load_workbook(path)
    cell = wb["Cover"]["B3"]
    cell.font = Font(color=BLUE)
    wb.save(path)

    report = audit_workbook(path)
    assert report.passed, report.violations


def test_audit_flags_named_range_with_multiple_destinations(tmp_path):
    path = _write(tmp_path)
    wb = load_workbook(path)
    # A name whose attr_text lists two comma-separated destinations resolves to
    # two cells — the audit wants exactly one.
    wb.defined_names["Goodwill"] = DefinedName("Goodwill", attr_text="'PPA'!$B$11,'PPA'!$B$10")
    wb.save(path)

    report = audit_workbook(path)
    assert not report.passed
    assert any("2 destinations" in v for v in report.violations)


def test_audit_report_summaries_reflect_pass_and_fail():
    assert "PASS" in AuditReport(cells_checked=3).summary()
    assert "VIOLATION" in AuditReport(violations=["x"], cells_checked=3).summary()
