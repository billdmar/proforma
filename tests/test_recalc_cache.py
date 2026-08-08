"""Tests for :func:`src.workbook.recalc_cache.cache_formula_values`.

The property under test (a recruiter-surface polish): after caching, a
no-recalc previewer — modelled here by ``load_workbook(data_only=True)`` —
returns numbers for the formula cells, while ``data_only=False`` still returns
the live formula strings. Both readings must coexist, and the cached numbers
must match the engine to the cent. Determinism (byte-identical rebuild) is also
asserted so the deliverable stays reproducible.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook
from src.workbook import ExcelWorkbookWriter, cache_formula_values

_PRECEDENTS_CSV = "data/curated/precedents_software.csv"

# Formula cells that must be blank before caching and numeric after (headline
# Cover tearsheet + representative Deal/PPA/Pro-Forma/Accretion formula cells).
_FORMULA_CELLS = [
    ("Cover", "B7"),  # equity purchase price
    ("Cover", "B8"),  # implied premium
    ("Cover", "B9"),  # goodwill created
    ("Deal & S&U", "B6"),  # equity purchase price build
    ("PPA", "B11"),  # goodwill (the plug)
    ("Pro Forma IS", "B14"),  # pro forma EPS, Y1
    ("Accretion-Dilution", "B6"),  # accretion/(dilution) Y1
]


@pytest.fixture(scope="module")
def _flagship():
    from src.flagship import build_flagship_bundle

    return build_flagship_bundle(_PRECEDENTS_CSV)


def _written(tmp_path, model, name="m.xlsx"):
    p = tmp_path / name
    ExcelWorkbookWriter().write(str(p), model)
    return p


def test_before_caching_previewers_see_blanks(tmp_path, _flagship):
    """Baseline: without caching, a data_only reader sees None (the bug)."""
    p = _written(tmp_path, _flagship)
    wb = load_workbook(p, data_only=True)
    for sheet, coord in _FORMULA_CELLS:
        assert wb[sheet][coord].value is None


def test_caching_populates_cached_values(tmp_path, _flagship):
    """After caching, data_only=True returns numbers for the formula cells."""
    p = _written(tmp_path, _flagship)
    cache_formula_values(str(p))
    wb = load_workbook(p, data_only=True)
    for sheet, coord in _FORMULA_CELLS:
        v = wb[sheet][coord].value
        assert isinstance(v, (int, float)), f"{sheet}!{coord} still blank: {v!r}"


def test_caching_preserves_live_formulas(tmp_path, _flagship):
    """After caching, data_only=False still returns the live formula strings."""
    p = _written(tmp_path, _flagship)
    cache_formula_values(str(p))
    wb = load_workbook(p, data_only=False)
    # The exact cross-sheet reference from the Cover tearsheet is unchanged.
    assert wb["Cover"]["B7"].value == "='Deal & S&U'!B6"
    for sheet, coord in _FORMULA_CELLS:
        assert wb[sheet][coord].data_type == "f", f"{sheet}!{coord} lost its formula"


def test_cached_values_match_engine_to_the_cent(tmp_path, _flagship):
    """Goodwill and implied premium cached values match the engine to the cent."""
    p = _written(tmp_path, _flagship)
    cache_formula_values(str(p))
    wb = load_workbook(p, data_only=True)

    ppa = _flagship.deal.ppa
    premium = _flagship.deal.consideration.implied_premium_pct

    assert wb["PPA"]["B11"].value == pytest.approx(ppa.goodwill, abs=0.01)
    assert wb["Cover"]["B9"].value == pytest.approx(ppa.goodwill, abs=0.01)
    assert wb["Cover"]["B8"].value == pytest.approx(premium, abs=0.01)
    # Sanity on the disclosed-deal magnitudes flagged by the task.
    assert wb["Cover"]["B9"].value == pytest.approx(25.4e9, abs=0.2e9)
    assert wb["Cover"]["B8"].value == pytest.approx(0.287, abs=0.01)


def test_caching_is_deterministic(tmp_path, _flagship):
    """Two write+cache runs produce a byte-identical file."""
    p1 = _written(tmp_path, _flagship, "a.xlsx")
    p2 = _written(tmp_path, _flagship, "b.xlsx")
    cache_formula_values(str(p1))
    cache_formula_values(str(p2))
    assert p1.read_bytes() == p2.read_bytes()
